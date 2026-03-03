#!/usr/bin/env python3
"""Build frontend JSON search indexes from source Excel workbooks.

Author: Dr Binghao Chai
Institute: University College London

This script reads the source Excel workbooks in `source/`, reads the HGNC
reference table from `database/`, and writes JSON files for the static search
page into `database/`.
"""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE_DIR = ROOT / "source"
DATABASE_DIR = ROOT / "database"
DIAGNOSIS_TO_FUSION_XLSX = SOURCE_DIR / "Diagnosis_to_gene_fusion_mapping.xlsx"
GENE_TO_DIAGNOSIS_XLSX = SOURCE_DIR / "Gene_to_diagnosis_mapping.xlsx"
GENE_TO_GENE_XLSX = SOURCE_DIR / "Gene_to_gene_fusion_mapping.xlsx"
HGNC_TSV = DATABASE_DIR / "hgnc_complete_set.txt"


@dataclass
class WorkbookTable:
    header: list[str]
    rows: list[list[str]]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_table(path: Path) -> WorkbookTable:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            raise ValueError(f"Workbook has no rows: {path}")
        header = [normalize_text(value) for value in header_row]
        rows = [[normalize_text(value) for value in row] for row in row_iter]
        return WorkbookTable(header=header, rows=rows)
    finally:
        workbook.close()


def split_pipe_values(raw_value: str) -> list[str]:
    if not raw_value or raw_value.lower() == "nan":
        return []
    return [part.strip() for part in raw_value.split("|") if part.strip()]


def build_alias_maps(hgnc_path: Path) -> tuple[dict[str, str], dict[str, list[str]]]:
    alias_lookup: dict[str, str] = {}
    canonical_to_aliases: dict[str, set[str]] = defaultdict(set)
    ambiguous_aliases: set[str] = set()
    alias_targets: dict[str, set[str]] = defaultdict(set)

    with hgnc_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            if normalize_text(row.get("status")) != "Approved":
                continue
            canonical = normalize_text(row.get("symbol"))
            if not canonical:
                continue
            alias_targets[canonical.upper()].add(canonical)
            for alias in split_pipe_values(normalize_text(row.get("prev_symbol"))):
                alias_targets[alias.upper()].add(canonical)
                canonical_to_aliases[canonical].add(alias)
            for alias in split_pipe_values(normalize_text(row.get("alias_symbol"))):
                alias_targets[alias.upper()].add(canonical)
                canonical_to_aliases[canonical].add(alias)

    for alias_key, targets in alias_targets.items():
        if len(targets) == 1:
            alias_lookup[alias_key] = next(iter(targets))
        else:
            ambiguous_aliases.add(alias_key)

    alias_lookup.update({key: value for key, value in alias_lookup.items() if key not in ambiguous_aliases})
    canonical_alias_map = {
        canonical: sorted({alias for alias in aliases if alias.upper() != canonical.upper()}, key=str.casefold)
        for canonical, aliases in canonical_to_aliases.items()
    }
    return alias_lookup, canonical_alias_map


def canonicalize_gene(symbol: str, alias_lookup: dict[str, str]) -> str:
    normalized = normalize_text(symbol)
    if not normalized:
        return ""
    return alias_lookup.get(normalized.upper(), normalized)


def extract_fusion_genes(fusion_value: str) -> tuple[str, str] | None:
    if not fusion_value or "::" not in fusion_value:
        return None
    left, right = fusion_value.split("::", 1)
    left = normalize_text(left)
    right = normalize_text(right)
    if not left or not right:
        return None
    return left, right


def sorted_unique(values: Iterable[str]) -> list[str]:
    return sorted({value for value in values if value}, key=str.casefold)


def build_indexes() -> dict[str, object]:
    full_alias_lookup, full_canonical_to_aliases = build_alias_maps(HGNC_TSV)

    diagnosis_table = load_table(DIAGNOSIS_TO_FUSION_XLSX)
    gene_diag_table = load_table(GENE_TO_DIAGNOSIS_XLSX)
    gene_gene_table = load_table(GENE_TO_GENE_XLSX)

    diagnosis_lookup: dict[str, dict[str, list[str] | str]] = {}
    diagnosis_key_lookup: dict[str, str] = {}
    fusion_lookup: dict[str, dict[str, object]] = {}
    gene_lookup: dict[str, dict[str, object]] = {}

    diagnosis_col_idx = diagnosis_table.header.index("diagnosis")
    fusion_col_indexes = [i for i, name in enumerate(diagnosis_table.header) if name.startswith("fusion_partner_")]

    for row in diagnosis_table.rows:
        diagnosis = normalize_text(row[diagnosis_col_idx]) if diagnosis_col_idx < len(row) else ""
        if not diagnosis:
            continue
        diagnosis_key_lookup[diagnosis.casefold()] = diagnosis
        fusions: list[str] = []
        genes: set[str] = set()
        for col_idx in fusion_col_indexes:
            if col_idx >= len(row):
                continue
            fusion_value = normalize_text(row[col_idx])
            genes_pair = extract_fusion_genes(fusion_value)
            if genes_pair is None:
                continue
            left = canonicalize_gene(genes_pair[0], full_alias_lookup)
            right = canonicalize_gene(genes_pair[1], full_alias_lookup)
            display_fusion = f"{left}::{right}"
            fusions.append(display_fusion)
            genes.update([left, right])
            fusion_key = "::".join(sorted([left, right], key=str.casefold))
            bucket = fusion_lookup.setdefault(
                fusion_key,
                {
                    "fusion": fusion_key,
                    "genes": sorted([left, right], key=str.casefold),
                    "diagnoses": set(),
                    "observed_fusions": set(),
                },
            )
            bucket["diagnoses"].add(diagnosis)
            bucket["observed_fusions"].add(display_fusion)
        diagnosis_lookup[diagnosis] = {
            "diagnosis": diagnosis,
            "genes": sorted(genes, key=str.casefold),
            "fusions": sorted_unique(fusions),
        }

    gene_col_idx = gene_diag_table.header.index("gene")
    diagnosis_value_indexes = [i for i, name in enumerate(gene_diag_table.header) if name.startswith("diagnosis_")]
    for row in gene_diag_table.rows:
        gene = canonicalize_gene(row[gene_col_idx] if gene_col_idx < len(row) else "", full_alias_lookup)
        if not gene:
            continue
        diagnoses = [normalize_text(row[i]) for i in diagnosis_value_indexes if i < len(row) and normalize_text(row[i])]
        gene_lookup.setdefault(
            gene,
            {
                "gene": gene,
                "diagnoses": [],
                "partner_genes": [],
                "aliases": full_canonical_to_aliases.get(gene, []),
            },
        )
        gene_lookup[gene]["diagnoses"] = sorted_unique([*gene_lookup[gene]["diagnoses"], *diagnoses])
        gene_lookup[gene]["aliases"] = full_canonical_to_aliases.get(gene, [])

    gene_col_idx = gene_gene_table.header.index("gene")
    partner_value_indexes = [i for i, name in enumerate(gene_gene_table.header) if name.startswith("partner_gene_")]
    for row in gene_gene_table.rows:
        gene = canonicalize_gene(row[gene_col_idx] if gene_col_idx < len(row) else "", full_alias_lookup)
        if not gene:
            continue
        partners = [
            canonicalize_gene(row[i], full_alias_lookup)
            for i in partner_value_indexes
            if i < len(row) and normalize_text(row[i])
        ]
        gene_lookup.setdefault(
            gene,
            {
                "gene": gene,
                "diagnoses": [],
                "partner_genes": [],
                "aliases": full_canonical_to_aliases.get(gene, []),
            },
        )
        gene_lookup[gene]["partner_genes"] = sorted_unique([*gene_lookup[gene]["partner_genes"], *partners])
        gene_lookup[gene]["aliases"] = full_canonical_to_aliases.get(gene, [])

    for diagnosis, record in diagnosis_lookup.items():
        for gene in record["genes"]:
            gene_lookup.setdefault(
                gene,
                {
                    "gene": gene,
                    "diagnoses": [],
                    "partner_genes": [],
                    "aliases": full_canonical_to_aliases.get(gene, []),
                },
            )
            gene_lookup[gene]["diagnoses"] = sorted_unique([*gene_lookup[gene]["diagnoses"], diagnosis])

    dataset_genes = set(gene_lookup)
    alias_lookup = {
        alias: canonical
        for alias, canonical in full_alias_lookup.items()
        if canonical in dataset_genes
    }
    canonical_to_aliases = {
        gene: sorted(full_canonical_to_aliases.get(gene, []), key=str.casefold)
        for gene in dataset_genes
    }
    for gene, record in gene_lookup.items():
        record["aliases"] = canonical_to_aliases.get(gene, [])

    gene_key_lookup = {gene.casefold(): gene for gene in gene_lookup}
    fusion_key_lookup = {fusion.casefold(): fusion for fusion in fusion_lookup}

    normalized_fusion_lookup = {}
    for fusion_key, record in fusion_lookup.items():
        normalized_fusion_lookup[fusion_key] = {
            "fusion": record["fusion"],
            "genes": record["genes"],
            "diagnoses": sorted(record["diagnoses"], key=str.casefold),
            "observed_fusions": sorted(record["observed_fusions"], key=str.casefold),
        }

    diagnosis_lookup_sorted = {
        diagnosis: {
            "diagnosis": diagnosis,
            "genes": sorted(record["genes"], key=str.casefold),
            "fusions": sorted(record["fusions"], key=str.casefold),
        }
        for diagnosis, record in sorted(diagnosis_lookup.items(), key=lambda item: item[0].casefold())
    }

    gene_lookup_sorted = {
        gene: {
            "gene": gene,
            "diagnoses": sorted(record["diagnoses"], key=str.casefold),
            "partner_genes": sorted(record["partner_genes"], key=str.casefold),
            "aliases": sorted(record["aliases"], key=str.casefold),
        }
        for gene, record in sorted(gene_lookup.items(), key=lambda item: item[0].casefold())
    }

    source_json = {
        "diagnosis_to_gene_fusion_mapping": [
            diagnosis_lookup_sorted[diagnosis]
            for diagnosis in sorted(diagnosis_lookup_sorted, key=str.casefold)
        ],
        "gene_to_diagnosis_mapping": [
            {
                "gene": gene,
                "diagnoses": gene_lookup_sorted[gene]["diagnoses"],
            }
            for gene in sorted(gene_lookup_sorted, key=str.casefold)
        ],
        "gene_to_gene_fusion_mapping": [
            {
                "gene": gene,
                "partner_genes": gene_lookup_sorted[gene]["partner_genes"],
            }
            for gene in sorted(gene_lookup_sorted, key=str.casefold)
        ],
    }

    search_index = {
        "meta": {
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
            "counts": {
                "genes": len(gene_lookup_sorted),
                "diagnoses": len(diagnosis_lookup_sorted),
                "fusions": len(normalized_fusion_lookup),
                "aliases": len(alias_lookup),
            },
        },
        "gene_lookup": gene_lookup_sorted,
        "gene_key_lookup": gene_key_lookup,
        "diagnosis_lookup": diagnosis_lookup_sorted,
        "diagnosis_key_lookup": diagnosis_key_lookup,
        "fusion_lookup": dict(sorted(normalized_fusion_lookup.items(), key=lambda item: item[0].casefold())),
        "fusion_key_lookup": fusion_key_lookup,
        "suggestions": {
            "genes": sorted(gene_lookup_sorted.keys(), key=str.casefold),
            "diagnoses": sorted(diagnosis_lookup_sorted.keys(), key=str.casefold),
            "fusions": sorted(normalized_fusion_lookup.keys(), key=str.casefold),
            "aliases": [
                {"alias": alias, "canonical": canonical}
                for alias, canonical in sorted(alias_lookup.items(), key=lambda item: item[0].casefold())
                if alias != canonical.upper()
            ],
        },
    }

    return {
        "alias_lookup": dict(sorted(alias_lookup.items(), key=lambda item: item[0].casefold())),
        "search_index": search_index,
        "source_json": source_json,
    }


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> int:
    DATABASE_DIR.mkdir(parents=True, exist_ok=True)
    indexes = build_indexes()

    write_json(DATABASE_DIR / "alias_lookup.json", indexes["alias_lookup"])
    write_json(DATABASE_DIR / "search_index.json", indexes["search_index"])
    write_json(
        DATABASE_DIR / "diagnosis_to_gene_fusion_mapping.json",
        indexes["source_json"]["diagnosis_to_gene_fusion_mapping"],
    )
    write_json(
        DATABASE_DIR / "gene_to_diagnosis_mapping.json",
        indexes["source_json"]["gene_to_diagnosis_mapping"],
    )
    write_json(
        DATABASE_DIR / "gene_to_gene_fusion_mapping.json",
        indexes["source_json"]["gene_to_gene_fusion_mapping"],
    )

    print("Build summary:")
    print(f"- alias_lookup: {DATABASE_DIR / 'alias_lookup.json'}")
    print(f"- search_index: {DATABASE_DIR / 'search_index.json'}")
    print(f"- diagnosis_json: {DATABASE_DIR / 'diagnosis_to_gene_fusion_mapping.json'}")
    print(f"- gene_to_diagnosis_json: {DATABASE_DIR / 'gene_to_diagnosis_mapping.json'}")
    print(f"- gene_to_gene_json: {DATABASE_DIR / 'gene_to_gene_fusion_mapping.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
